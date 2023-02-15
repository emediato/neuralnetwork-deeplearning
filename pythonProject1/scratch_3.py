from openpyxl import load_workbook

wb = load_workbook(filename='Accounts.xlsx')
sheet = wb['sheet_name']

names = []

for i in range(1, sheet.max_row + 1):
    for j in range(sheet.max_column):
        if sheet[i][j].value:
            names.append(sheet[i][j].value)

#Then just check whether a username is in the list

if username in names: